#! /usr/bin/python3

# Multipurpose batch image processing application
# Copyright (c) Manjunathan1998
# https://github.com/Manjunathan1998/Multipurpose-batch-image-processing-application


from keras.preprocessing.image import ImageDataGenerator, array_to_img, img_to_array, load_img
from ttkthemes import themed_tk as tk
from hurry.filesize import size
import matplotlib.pyplot as plt
from tkinter import filedialog
import tkinter.messagebox
from PIL import ImageTk
from tkinter import ttk
import PIL.ImageFont
import PIL.ImageDraw
from tkinter import *
from math import *
import pandas as pd
import numpy as np
import openpyxl
import glob
import cv2
import os
#############################################################################################################
window = tk.ThemedTk()
window.get_themes()
window.set_theme('black') 
window.title("Batch Image Processing")
window.geometry("1024x720")
window.iconbitmap(r'Support/icons/software.ico')

def icon(name):
	path = "Support/icons/"+str(name)
	return path

global CheckVar1, CheckVar2, mark_1, mark_2, mark_3, mark_4
global brightness_var, contrast_var
brightness_var = DoubleVar() 
contrast_var = DoubleVar()
CheckVar1 = IntVar()
CheckVar2 = IntVar()
mark_1 = IntVar()
mark_2 = IntVar()
mark_3 = IntVar()
mark_4 = IntVar()
global current_stack, get_dummy
current_stack = {"stack": "", "Folder": "", "Adjustments": ""}
def exit():
	window.destroy()
def message_box(title, message):
	tkinter.messagebox.showinfo(title, message)
def db():
	global new_color, new_height, new_width
	df = pd.read_excel('Support/log.xlsx', index_col = False)
	new_color =  df.iloc[0,0]
	new_height = df.iloc[1,0]
	new_width = df.iloc[2,0]
db()
###########################################################################################
def operations(op_id, image_path):
	if (op_id == 1):
		df = pd.read_excel('Support/log.xlsx')
		bright = df.iloc[0, 1]
		contra = df.iloc[1, 1]
		message_box('Building elements', 'Brightness = '+str(bright)+' Contrast = '+str(contra))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				img = cv2.imread(f_img)
				img_conver = cv2.convertScaleAbs(img, alpha=contra, beta=bright)
				cv2.imwrite(f_img, img_conver)
			message_box("Completed", "Corrections applied")
		except:
			message_box("Error", "Something went wrong")
	if (op_id == 2):
		df = pd.read_excel('Support/log.xlsx')
		height_resize = df.iloc[0, 2]
		width_resize = df.iloc[1, 2]
		height_resize = int(height_resize)
		width_resize = int(width_resize)
		message_box('Building elements', 'Height = '+str(height_resize)+' Width = '+str(width_resize))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				img = cv2.imread(f_img)
				img_conver = cv2.resize(img , (round(width_resize), round(height_resize)))
				cv2.imwrite(f_img, img_conver)
			message_box("Completed", "Resized successfully")
		except:
			message_box("Error", "Something went wrong")
	if (op_id == 3):
		df = pd.read_excel('Support/log.xlsx')
		orient_text = df.iloc[0, 3]
		message_box('Building elements', 'Orientation = '+str(orient_text))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				if(orient_text == "Rotate by 90° cw"):
					img = cv2.imread(f_img)
					img_conver = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
					cv2.imwrite(f_img, img_conver)
				elif(orient_text == "Rotate by 180° cw"):
					img = cv2.imread(f_img)
					img_conver = cv2.rotate(img, cv2.ROTATE_180)
					cv2.imwrite(f_img, img_conver)
				elif(orient_text == "Rotate by 270° cw"):
					img = cv2.imread(f_img)
					img_conver = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
					cv2.imwrite(f_img, img_conver)
			message_box("Completed", "Orientation successful")
		except:
			message_box("Error", "Something went wrong")
	if (op_id == 4):
		df = pd.read_excel("Support/log.xlsx")
		text = df.iloc[0,4]
		type_text =df.iloc[1,4]
		i = 0
		f = image_path
		try:
			for filename in os.listdir(f):
				f_img = f+'/'+filename
				my_dest =text+str(i)+str('.')+type_text
				my_dest =f+'/'+ my_dest
				os.rename(f_img, my_dest)
				i += 1
			message_box("Completed", "Renaming successful")
		except:
			message_box("Error", "Something went wrong")
	if (op_id == 5):
		def verify_alpha_channel(frame):
			try:
				frame.shape[3] 
			except IndexError:
				frame = cv2.cvtColor(frame, cv2.COLOR_BGR2BGRA)
				return frame
		df = pd.read_excel("Support/log.xlsx")
		text_filter = df.iloc[0,5]
		message_box('Building elements', 'Filter = '+str(text_filter))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				def color_process(color_b,color_g, color_r):
					img = cv2.imread(f_img)
					frame = verify_alpha_channel(img)
					frame_h, frame_w, frame_c = frame.shape
					sepia_bgra = (color_b, color_g, color_r, 1)
					overlay = np.full((frame_h, frame_w, 4), sepia_bgra, dtype='uint8')
					img_conver = cv2.addWeighted(overlay, 0.5, frame, 1.0, 0, frame)
					cv2.imwrite(f_img, img_conver)
				if text_filter == "Grayscale":
					img = cv2.imread(f_img)
					img_conver = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
					cv2.imwrite(f_img, img_conver)
				elif text_filter == "Sepia":
					img = cv2.imread(f_img)
					frame = verify_alpha_channel(img)
					frame_h, frame_w, frame_c = frame.shape
					sepia_bgra = (20, 66, 112, 1)
					overlay = np.full((frame_h, frame_w, 4), sepia_bgra, dtype='uint8')
					img_conver = cv2.addWeighted(overlay, 0.5, frame, 1.0, 0, frame)
					cv2.imwrite(f_img, img_conver)
				elif text_filter == "Threshold":
					img = cv2.imread(f_img)
					gray = cv2.cvtColor(np.float32(img), cv2.COLOR_BGR2GRAY)
					_,mask = cv2.threshold(np.float32(gray), 120,255,cv2.THRESH_BINARY)
					mask = cv2.cvtColor(np.float32(mask), cv2.COLOR_GRAY2BGRA)
					img_conver = cv2.cvtColor(mask, cv2.COLOR_BGRA2BGR)
					cv2.imwrite(f_img, img_conver)
				elif text_filter == "Invert":
					img = cv2.imread(f_img)
					img_conver = cv2.bitwise_not(img)
					cv2.imwrite(f_img, img_conver)
				elif text_filter == "Color Overlay 'R'":
					color_process(0,0,255)
				elif text_filter == "Color Overlay 'G'":
					color_process(0,255,0)
				elif text_filter == "Color Overlay 'B'":
					color_process(255,0,0)
			message_box("Completed", "Filter applied successfully")
		except:
			message_box("Error", "Something went wrong")	
	if (op_id == 6):
		df = pd.read_excel("Support/log.xlsx")
		text_colormap = df.iloc[0,6]
		message_box('Building elements', 'Colormap = '+str(text_colormap))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				if text_colormap == "viridis":
					img = cv2.imread(f_img,0)
					colormap = plt.get_cmap("viridis")
					heatmap = (colormap(img) * 2**16).astype(np.uint16)[:,:,:3]
					img_conver = cv2.cvtColor(heatmap, cv2.COLOR_RGB2BGR)
					cv2.imwrite(f_img, img_conver)
				elif text_colormap == "plasma":
					img = cv2.imread(f_img,0)
					colormap = plt.get_cmap("plasma")
					heatmap = (colormap(img) * 2**16).astype(np.uint16)[:,:,:3]
					img_conver = cv2.cvtColor(heatmap, cv2.COLOR_RGB2BGR)
					cv2.imwrite(f_img, img_conver)
				elif text_colormap == "inferno":
					img = cv2.imread(f_img,0)
					colormap = plt.get_cmap("inferno")
					heatmap = (colormap(img) * 2**16).astype(np.uint16)[:,:,:3]
					img_conver = cv2.cvtColor(heatmap, cv2.COLOR_RGB2BGR)
					cv2.imwrite(f_img, img_conver)
				elif text_colormap == "magma":
					img = cv2.imread(f_img,0)
					colormap = plt.get_cmap("magma")
					heatmap = (colormap(img) * 2**16).astype(np.uint16)[:,:,:3]
					img_conver = cv2.cvtColor(heatmap, cv2.COLOR_RGB2BGR)
					cv2.imwrite(f_img, img_conver)
				elif text_colormap == "cividis":
					img = cv2.imread(f_img,0)
					colormap = plt.get_cmap("cividis")
					heatmap = (colormap(img) * 2**16).astype(np.uint16)[:,:,:3]
					img_conver = cv2.cvtColor(heatmap, cv2.COLOR_RGB2BGR)
					cv2.imwrite(f_img, img_conver)
			message_box("Completed", "Colormap applied successfully")
		except:
			message_box("Error", "Something went wrong")
	if (op_id == 7):
		df = pd.read_excel("Support/log.xlsx")
		text_blur = df.iloc[0,7]
		message_box('Building elements', 'Blur = '+str(text_blur))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				blur_state = (int(text_blur), int(text_blur))
				img = cv2.imread(f_img)
				img_conver = cv2.blur(img, blur_state)
				cv2.imwrite(f_img, img_conver)
			message_box("Completed", "Successfully applied")
		except:
			message_box("Error", "Something went wrong")
	if (op_id == 8):
		df = pd.read_excel("Support/log.xlsx")
		text_flip = df.iloc[0,8]
		message_box('Building elements', 'Flip = '+str(text_flip))
		try:
			f = image_path
			for file in os.listdir(f):
				f_img = f+'/'+file
				if text_flip == "Flip vertical":
					img = cv2.imread(f_img)
					img_conver = cv2.flip(img, 0)
					cv2.imwrite(f_img, img_conver)
				elif text_flip == "Flip horizontal":
					img = cv2.imread(f_img)
					img_conver = cv2.flip(img, 1)
					cv2.imwrite(f_img, img_conver)
				elif text_flip == "Flip horizontal & vertical":
					img = cv2.imread(f_img)
					img_conver = cv2.flip(img, -1)
					cv2.imwrite(f_img, img_conver)
			message_box("Completed", "Flipping successful")
		except:
			message_box("Error","Something went wrong")
############################################################################################
def normal():
	os.remove('Work_Dir/img.png')
	get_dummy('Support/dummy.png', 'Work_Dir/img.png')
	disp_image('Work_Dir/img.png')
	os.remove('Support/dummy.png')
def get_dummy(path_of_origianl, path_to_save):
	original_img = cv2.imread(path_of_origianl)
	cv2.imwrite(path_to_save, original_img)
def increase_brightness(alphavalue, betavalue):
	get_dummy("Support/dummy.png", "Work_Dir/img.png")
	frame_dummy = cv2.imread("Work_Dir/img.png")
	img_dummy = cv2.convertScaleAbs(frame_dummy, alpha=alphavalue, beta=betavalue)
	cv2.imwrite("Work_Dir/img.png", img_dummy)
	disp_image("Work_Dir/img.png")
def img_corrections():
	def apply_mask():
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['B1']= 'B,C'
		ws['B2'] = round(b_value)
		ws['B3'] = round(c_value)
		book.save('Support/log.xlsx')
		message_box('Done', 'Mask setting are saved')
	def cancel_adj():
		brightness_label.destroy()
		contrast_label.destroy()
		s1.destroy()
		s2.destroy()
		mask_adj_btn.destroy()
		apply_adj_btn.destroy()
		cancel_adj_btn.destroy()
		reset_adj_btn.destroy()
		normal()
		b_value_label = ttk.Label(main_window, text = "")
		b_value_label.place(relx = 0.93, rely = 0.63, relheight = 0.03, relwidth = 0.05)
		c_value_label = ttk.Label(main_window, text = "")
		c_value_label.place(relx = 0.93, rely = 0.73, relheight = 0.03, relwidth = 0.05)
	def apply_action():
		global b_value_label
		global c_value_label
		global b_value, c_value
		b_value = s1.get()
		b_value= float(b_value)
		b_value = round(b_value)
		b_value_label = ttk.Label(main_window, text = str(b_value))
		b_value_label.place(relx = 0.93, rely = 0.63, relheight = 0.03, relwidth = 0.05)
		c_value = s2.get()
		c_value= float(c_value)
		c_value = round(c_value)
		c_value_label = ttk.Label(main_window, text = str(c_value))
		c_value_label.place(relx = 0.93, rely = 0.73, relheight = 0.03, relwidth = 0.05)
		increase_brightness(c_value, b_value)
	def reset_adj():
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		s1.set(1)
		s2.set(1)
		apply_action()
		src_adj = cv2.imread("Work_Dir/img.png")
		cv2.imwrite("Work_Dir/img.png", src_adj)
		disp_image("Work_Dir/img.png")
	if (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		def apply_bulk_corrections():
			operations(1, current_stack['Folder'])		
		def cancel_bulk_corrections():
			folder_label.destroy()
			folder_name_label.destroy()
			apply_corrections_btn.destroy()
			cancel_corrections_btn.destroy()
		folder_label = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		apply_corrections_btn = ttk.Button(main_window, text = "Apply Mask", command = apply_bulk_corrections)
		apply_corrections_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_corrections_btn = ttk.Button(main_window, text = "Cancel", command = cancel_bulk_corrections)
		cancel_corrections_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
	elif (current_stack["stack"] == "Single"):
		global brightness_label, contrast_label, s1, s2
		current_stack["Adjustments"] = "True"
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		s1 = ttk.Scale(main_window,from_ = 1, to = 50, orient = HORIZONTAL)
		s1.set(1) 
		s1.place(relx = 0.86, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		s2 = ttk.Scale(main_window,from_ = 1, to = 5, orient = HORIZONTAL)
		s2.set(1) 
		s2.place(relx = 0.86, rely = 0.7, relheight = 0.03, relwidth = 0.1)
		brightness_label = ttk.Label(main_window, text = "Brightness")
		brightness_label.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.08)
		contrast_label = ttk.Label(main_window, text = "Contrast")
		contrast_label.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.08)
		mask_adj_btn = ttk.Button(main_window, text = "Set Mask", command = apply_mask)
		mask_adj_btn.place(relx = 0.76, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		apply_adj_btn = ttk.Button(main_window, text = "Apply",command =  apply_action)
		apply_adj_btn.place(relx = 0.89, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		cancel_adj_btn = ttk.Button(main_window, text = "Cancel", command = cancel_adj)
		cancel_adj_btn.place(relx = 0.76, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		reset_adj_btn = ttk.Button(main_window, text = "Reset", command = reset_adj)
		reset_adj_btn.place(relx = 0.89, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
	elif (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
###########################################################################################
def resizer():
	def cancel_res():
		folder_label.destroy()
		folder_name_label.destroy()
		height_label.destroy()
		width_label.destroy()
		height_entry_resize.destroy()
		width_entry_resize.destroy()
		apply_res_btn.destroy()
		cancel_res_btn.destroy()
	def apply_res():
		height_val_re = height_val_resize.get()
		width_val_re = width_val_resize.get()
		height_re = float(height_val_re)
		width_re = float(width_val_re)
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['C1']= 'H,W'
		ws['C2'] = round(height_re)
		ws['C3'] = round(width_re)
		book.save('Support/log.xlsx')
		operations(2, current_stack["Folder"])
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		message_box("Error", "Resizer works with bulk images")
	elif (current_stack["stack"] == "Bulk"):
		df = pd.read_excel('Support/log.xlsx')
		height = df.iloc[0,2]
		width = df.iloc[1,2]
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		folder_label = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		height_label = ttk.Label(main_window, text = "Height: ")
		height_label.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		width_label = ttk.Label(main_window, text = "Width: ")
		width_label.place(relx = 0.76, rely = 0.75, relheight = 0.03, relwidth = 0.09)
		apply_res_btn = ttk.Button(main_window, text = "Apply", command = apply_res)
		apply_res_btn.place(relx = 0.77, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		cancel_res_btn = ttk.Button(main_window, text = "Cancel", command = cancel_res)
		cancel_res_btn.place(relx = 0.88, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		global height_val_resize
		height_val_resize = StringVar()
		height_entry_resize = ttk.Entry(main_window,textvariable = height_val_resize)
		height_entry_resize.insert(0, height)
		height_entry_resize.place(relx = 0.81, rely = 0.7, relheight = 0.03, relwidth = 0.1)
		global width_val_resize
		width_val_resize = StringVar()
		width_entry_resize = ttk.Entry(main_window,textvariable = width_val_resize)
		width_entry_resize.insert(0, width)
		width_entry_resize.place(relx = 0.81, rely = 0.75, relheight = 0.03, relwidth = 0.1)
###########################################################################################
def image_orient():
	def apply_orient():
		orient_val = variables_imageorient.get()
		if (orient_val == ""):
			message_box("Error", "Select orientation")
		elif (orient_val == "Rotate by 90° cw"):
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			src_orient = cv2.imread("Work_Dir/img.png")
			image_orient_dummy = cv2.rotate(src_orient, cv2.ROTATE_90_CLOCKWISE)
			cv2.imwrite("Work_Dir/img.png", image_orient_dummy)
			disp_image("Work_Dir/img.png")
		elif (orient_val == "Rotate by 180° cw"):
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			src_orient = cv2.imread("Work_Dir/img.png")
			image_orient_dummy = cv2.rotate(src_orient, cv2.ROTATE_180)
			cv2.imwrite("Work_Dir/img.png", image_orient_dummy)
			disp_image("Work_Dir/img.png")
		elif (orient_val == "Rotate by 270° cw"):
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			src_orient = cv2.imread("Work_Dir/img.png")
			image_orient_dummy = cv2.rotate(src_orient, cv2.ROTATE_90_COUNTERCLOCKWISE)
			cv2.imwrite("Work_Dir/img.png", image_orient_dummy)
			disp_image("Work_Dir/img.png")
	def apply_mask_orient():
		orient_val = variables_imageorient.get()
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['D1']= 'Rotation'
		ws['D2'] = orient_val
		book.save('Support/log.xlsx')
		message_box('Done', 'Mask setting are saved')
	def reset_orient():
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		src_orient = cv2.imread("Work_Dir/img.png")
		cv2.imwrite("Work_Dir/img.png", src_orient)
		disp_image("Work_Dir/img.png")
	def cancel_orient():
		orient_label.destroy()
		combo_box_orient.destroy()
		mask_ore_btn.destroy()
		apply_ore_btn.destroy()
		cancel_ore_btn.destroy()
		reset_ore_btn.destroy()
		#reset_orient()
		normal()
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		current_stack["Adjustments"] = "True"
		orient_datatypes = ["Rotate by 90° cw", "Rotate by 180° cw", "Rotate by 270° cw"]
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		orient_label = ttk.Label(main_window, text = "Orientation type: ")
		orient_label.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		global variables_imageorient
		variables_imageorient = StringVar()
		combo_box_orient = ttk.Combobox(main_window,textvariable =variables_imageorient,font = ('arial',8), values = orient_datatypes)
		combo_box_orient.place(relx = 0.86, rely = 0.6, relheight = 0.03, relwidth = 0.12)
		mask_ore_btn = ttk.Button(main_window, text = "Set Mask", command = apply_mask_orient)
		mask_ore_btn.place(relx = 0.76, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		apply_ore_btn = ttk.Button(main_window, text = "Apply", command = apply_orient)
		apply_ore_btn.place(relx = 0.89, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		cancel_ore_btn = ttk.Button(main_window, text = "Cancel", command = cancel_orient)
		cancel_ore_btn.place(relx = 0.76, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		reset_ore_btn = ttk.Button(main_window, text = "Reset", command = reset_orient)
		reset_ore_btn.place(relx = 0.89, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
	elif (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		def apply_bulk_orientation():
			operations(3, current_stack['Folder'])		
		def cancel_bulk_orientation():
			folder_label_orient.destroy()
			folder_name_label_orient.destroy()
			apply_orient_btn.destroy()
			cancel_orient_btn.destroy()
		folder_label_orient = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label_orient.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label_orient = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label_orient.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		apply_orient_btn = ttk.Button(main_window, text = "Apply Mask", command = apply_bulk_orientation)
		apply_orient_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_orient_btn = ttk.Button(main_window, text = "Cancel", command = cancel_bulk_orientation)
		cancel_orient_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
###########################################################################################
def rename_images():
	df = pd.read_excel("Support/log.xlsx")
	text_1 = df.iloc[0,4]
	name_1 =df.iloc[1,4]
	text_name = text_1+'.'+name_1
	def apply_rename():
		text_name = text_val.get()
		text_name_split = text_name.split('.')
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['E1']= 'H,W'
		ws['E2'] = text_name_split[0]
		ws['E3'] = text_name_split[1]
		book.save('Support/log.xlsx')
		operations(4, current_stack["Folder"])
	def rename_cancel():
		folder_label_rename.destroy()
		folder_name_label_rename.destroy()
		name_rename.destroy()
		text_entry.destroy()
		apply_rename.destroy()
		cancel_rename.destroy()
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		message_box("Error", "Renaming works with bulk images")
	elif (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		folder_label_rename = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label_rename.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label_rename = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label_rename.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		name_rename = ttk.Label(main_window, text = "Text: ")
		name_rename.place(relx = 0.76, rely = 0.65, relheight = 0.03, relwidth = 0.1)
		global text_val
		text_val = StringVar()
		text_entry = ttk.Entry(main_window,textvariable = text_val)
		text_entry.insert(0, text_name)
		text_entry.place(relx = 0.81, rely = 0.65, relheight = 0.03, relwidth = 0.15)
		apply_rename = ttk.Button(main_window, text = "Rename all", command = apply_rename)
		apply_rename.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_rename = ttk.Button(main_window, text = "Cancel", command = rename_cancel)
		cancel_rename.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
###########################################################################################
def bitwise():
	bitwise_datatypes = ["And", "Or", "Xor", "Not"]
	message = ("Note: For NOT operation\nImg 1 is considered and Img 2 is neglected")	
	def open_img1():
		global filename_img1
		filename_img1 = filedialog.askopenfilename()
		global get_img_1
		get_img_1 = filename_img1
		get_img_1_base = os.path.basename(get_img_1)
		label_path =ttk.Label(main_window, text = get_img_1_base)
		label_path.place(relx = 0.81, rely = 0.6, relheight = 0.03, relwidth = 0.1)
	def open_img2():
		global filename_img2
		filename_img2 = filedialog.askopenfilename()
		global get_img_2
		get_img_2 = filename_img2
		get_img_2_base = os.path.basename(get_img_2)
		label_path =ttk.Label(main_window, text = get_img_2_base)
		label_path.place(relx = 0.81, rely = 0.65, relheight = 0.03, relwidth = 0.1)
	def apply_bitwise():
		btwise_var = variables_bitwise.get()
		try:
			get_img_1_state = get_img_1
			get_img_2_state = get_img_2
		except:
			message_box("Error", "Details are missing")
		if (btwise_var == "") and (get_img_2_state == "") and (get_img_2_state == ""):
			message_box("Error", "Details are missing")
		else:
			try:
				if (btwise_var == "And"):
					bit_1 = cv2.imread(get_img_1)
					bit_2 = cv2.imread(get_img_2)
					bit_1_re = cv2.resize(bit_1  , (630 , 500))
					bit_2_re = cv2.resize(bit_2  , (630 , 500))
					bit_and = cv2.bitwise_and(bit_1_re, bit_2_re)
					cv2.imwrite("Work_Dir/img.png", bit_and)
					cv2.imwrite("Work_Dir/original_copy.png", bit_and)
					disp_image("Work_Dir/img.png")
				elif (btwise_var == "Or"):
					bit_1 = cv2.imread(get_img_1)
					bit_2 = cv2.imread(get_img_2)
					bit_1_re = cv2.resize(bit_1  , (630 , 500))
					bit_2_re = cv2.resize(bit_2  , (630 , 500))
					bit_or = cv2.bitwise_or(bit_1_re, bit_2_re)
					cv2.imwrite("Work_Dir/img.png", bit_or)
					cv2.imwrite("Work_Dir/original_copy.png", bit_or)
					disp_image("Work_Dir/img.png")
				elif (btwise_var == "Xor"):
					bit_1 = cv2.imread(get_img_1)
					bit_2 = cv2.imread(get_img_2)
					bit_1_re = cv2.resize(bit_1  , (630 , 500))
					bit_2_re = cv2.resize(bit_2  , (630 , 500))
					bit_xor = cv2.bitwise_xor(bit_1_re, bit_2_re)
					cv2.imwrite("Work_Dir/img.png", bit_xor)
					cv2.imwrite("Work_Dir/original_copy.png", bit_xor)
					disp_image("Work_Dir/img.png")
				elif (btwise_var == "Not"):
					bit_1 = cv2.imread(get_img_1)
					bit_1_re = cv2.resize(bit_1  , (630 , 500))
					bit_not = cv2.bitwise_not(bit_1_re)
					cv2.imwrite("Work_Dir/img.png", bit_not)
					cv2.imwrite("Work_Dir/original_copy.png", bit_not)
					disp_image("Work_Dir/img.png")
				else:
					message_box("Error", "Unknown operation")
			except:
				message_box("Error", "Something Went Wrong")
	def cancel_bit():
		image_1.destroy()
		image_2.destroy()
		op.destroy()
		browse_btn_1.destroy()
		browse_btn_2.destroy()
		combo_box_bitwise.destroy()
		apply_bit.destroy()
		cancel_bit.destroy()
		message_label.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	message_box("Warning", "Closing work area")
	disp_label = ttk.Label(main_window)
	disp_label.place(relx = 0.09, rely = 0.15, relheight = 0.7, relwidth = 0.6)
	cancel()
	current_stack["stack"] = "Single"
	current_stack["Adjustments"] = "True"
	text_head_adj = ttk.Label(main_window)
	text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	image_1 = ttk.Label(main_window, text = "Image 1: ")
	image_1.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
	image_2 = ttk.Label(main_window, text = "Image 2: ")
	image_2.place(relx = 0.76, rely = 0.65, relheight = 0.03, relwidth = 0.1)
	op = ttk.Label(main_window, text = "Operation: ")
	op.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.1)
	browse_btn_1 = ttk.Button(main_window,text = "Browse",command = open_img1)
	global variables_bitwise
	variables_bitwise = StringVar()
	combo_box_bitwise = ttk.Combobox(main_window,textvariable =variables_bitwise,font = ('arial',8), values = bitwise_datatypes)
	combo_box_bitwise.place(relx = 0.85, rely = 0.7, relheight = 0.03, relwidth = 0.1)
	browse_btn_1.place(relx = 0.92, rely = 0.6, relheight = 0.03, relwidth = 0.06)
	browse_btn_2 = ttk.Button(main_window,text = "Browse", command = open_img2)
	browse_btn_2.place(relx = 0.92, rely = 0.65, relheight = 0.03, relwidth = 0.06)
	apply_bit = ttk.Button(main_window, text = "Apply", command = apply_bitwise)
	apply_bit.place(relx = 0.77, rely = 0.79, relheight = 0.03, relwidth = 0.09)
	cancel_bit = ttk.Button(main_window, text = "Cancel",command = cancel_bit)
	cancel_bit.place(relx = 0.88, rely = 0.79, relheight = 0.03, relwidth = 0.09)
	message_label = Label(main_window, text = message)
	message_label.place(relx = 0.75, rely = 0.83, relheight = 0.05, relwidth = 0.24)
###########################################################################################
def data_augment():
	def browse_dir():
		filename_dir_cpy = filedialog.askdirectory()
		global get_dir_cpy
		get_dir_cpy = filename_dir_cpy
		folder_label = ttk.Label(main_window, text = (os.path.basename(get_dir_cpy)))
		folder_label.place(relx = 0.82, rely = 0.575, relheight = 0.03, relwidth = 0.1)
		current_stack["Folder"] =  get_dir_cpy
		return get_dir_cpy
	def save_dt():
		get_cpy = data_cpy.get()
		if (get_cpy == ""):
			message_box("Error", "Copy details are missing")
		elif(get_dir_cpy == ""):
			message_box("Error", "Directory details are missing")
		else:
			try:
				datagen = ImageDataGenerator(
	                   rotation_range = 50,
	                   width_shift_range = 0.2,
	                   height_shift_range = 0.2,
	                   shear_range = 0.2,
	                   zoom_range = 0.2,
	                   horizontal_flip = True,
	                   fill_mode = 'nearest')
				img = load_img("Work_Dir/img.png")
				x = img_to_array(img)
				x = x.reshape((1,) + x.shape) 
				i = 0 
				for batch in datagen.flow(x, batch_size = 1, save_to_dir = get_dir_cpy, save_prefix = 'img', save_format = 'jpeg'):
					i += 1
					if i > int(get_cpy):
						message_box("Success", "Images saved in your directory")
						break
			except:
				message_box("Error", "Something went wrong")
	def cancel_cpy():
		cpy_entry.destroy()
		browse_data.destroy()
		save_data.destroy()
		cancel_data.destroy()
		folder_label.destroy()
		adj_para("", "n")
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		adj_para("\n Folder:\n\n Copies:", 'nw')
		folder_label = ttk.Label(main_window, text = "Choose Folder")
		folder_label.place(relx = 0.82, rely = 0.575, relheight = 0.03, relwidth = 0.1)
		global data_cpy
		data_cpy = StringVar()
		cpy_entry = ttk.Entry(main_window,textvariable = data_cpy)
		cpy_entry.place(relx = 0.82, rely = 0.63, relheight = 0.03, relwidth = 0.1)
		browse_data = ttk.Button(main_window, text = "Browse", command = browse_dir)
		browse_data.place(relx = 0.925, rely = 0.575, relheight = 0.03, relwidth = 0.06)
		save_data = ttk.Button(main_window, text = "Brew", command = save_dt)
		save_data.place(relx = 0.8, rely = 0.7, relheight = 0.03, relwidth = 0.06)
		cancel_data = ttk.Button(main_window, text = "Cancel", command = cancel_cpy)
		cancel_data.place(relx = 0.87, rely = 0.7, relheight = 0.03, relwidth = 0.06)
	elif (current_stack["stack"] == "Bulk"):
		message_box("Error", "Bulk selection")
###########################################################################################
def histogram():
	if (current_stack["stack"] == "Single"):
		try:
			img_hist = cv2.imread("Work_Dir/img.png")
			b,g,r = cv2.split(img_hist)
			plt.title("Histogram")	
			plt.hist(b.ravel(), 256, [0,256])
			plt.hist(g.ravel(), 256, [0,256])
			plt.hist(r.ravel(), 256, [0,256])
			plt.xlabel("RGB")
			plt.ylabel("Frequency")
			plt.show()
		except:
			pass
	elif(current_stack["stack"] == "Bulk"):
		message_box("Error", "Can't apply histogram for bulk images")
	elif (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	
###########################################################################################
def filter_tools():
	def verify_alpha_channel(frame):
		try:
			frame.shape[3] 
		except IndexError:
			frame = cv2.cvtColor(frame, cv2.COLOR_BGR2BGRA)
			return frame
	def color_process(color_b,color_g, color_r):
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		frame_dummy = cv2.imread("Work_Dir/img.png")
		frame = verify_alpha_channel(frame_dummy)
		frame_h, frame_w, frame_c = frame.shape
		sepia_bgra = (color_b, color_g, color_r, 1)
		overlay = np.full((frame_h, frame_w, 4), sepia_bgra, dtype='uint8')
		img_dummy = cv2.addWeighted(overlay, 0.5, frame, 1.0, 0, frame)
		cv2.imwrite("Work_Dir/img.png", img_dummy)
		disp_image("Work_Dir/img.png")
	filter_datatypes = ["Grayscale", "Sepia", "Threshold", "Invert", "Color Overlay 'R'", "Color Overlay 'G'", "Color Overlay 'B'"]
	current_stack["Adjustments"] = "True"
	text_head_adj = ttk.Label(main_window)
	text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	def cancel_filter():
		filter_selection.destroy()
		combo_box_filter.destroy()
		mask_fil_btn.destroy()
		apply_fil_btn.destroy()
		cancel_fil_btn.destroy()
		reset_fil_btn.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		normal()
	def reset_filter():
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		src_filter = cv2.imread("Work_Dir/img.png")
		cv2.imwrite("Work_Dir/img.png", src_filter)
		disp_image("Work_Dir/img.png")
	def apply_filters():
		filter_apply = variables_filters.get()
		if filter_apply == "":
			message_box("Error", "No filter is selected")
		elif filter_apply == "Grayscale":
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			frame_dummy = cv2.imread("Work_Dir/img.png")
			img_dummy = cv2.cvtColor(frame_dummy, cv2.COLOR_BGR2GRAY) 
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		elif filter_apply == "Sepia":
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			frame_dummy = cv2.imread("Work_Dir/img.png")
			frame = verify_alpha_channel(frame_dummy)
			frame_h, frame_w, frame_c = frame.shape
			sepia_bgra = (20, 66, 112, 1)
			overlay = np.full((frame_h, frame_w, 4), sepia_bgra, dtype='uint8')
			img_dummy = cv2.addWeighted(overlay, 0.5, frame, 1.0, 0, frame)
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		elif filter_apply == "Threshold":
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			frame_dummy = cv2.imread("Work_Dir/img.png")
			gray = cv2.cvtColor(np.float32(frame_dummy), cv2.COLOR_BGR2GRAY)
			_,mask = cv2.threshold(np.float32(gray), 120,255,cv2.THRESH_BINARY)
			mask = cv2.cvtColor(np.float32(mask), cv2.COLOR_GRAY2BGRA)
			img_dummy = cv2.cvtColor(mask, cv2.COLOR_BGRA2BGR)
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		elif filter_apply == "Invert":
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			frame_dummy = cv2.imread("Work_Dir/img.png")
			img_dummy = cv2.bitwise_not(frame_dummy)
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		elif filter_apply == "Color Overlay 'R'":
			color_process(0,0,255)
		elif filter_apply == "Color Overlay 'G'":
			color_process(0,255,0)
		elif filter_apply == "Color Overlay 'B'":
			color_process(255,0,0)
	def set_filtermask():
		filter_apply = variables_filters.get()
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['F1']= 'Filter'
		ws['F2'] = filter_apply
		book.save('Support/log.xlsx')
		message_box("Done", "Mask setting are saved")
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		filter_selection = ttk.Label(main_window, text = "Select Filter ")
		filter_selection.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		global variables_filters
		variables_filters = StringVar()
		combo_box_filter = ttk.Combobox(main_window,textvariable =variables_filters,font = ('arial',8), values = filter_datatypes)
		combo_box_filter.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		mask_fil_btn = ttk.Button(main_window, text = "Set Mask", command = set_filtermask)
		mask_fil_btn.place(relx = 0.76, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		apply_fil_btn = ttk.Button(main_window, text = "Apply", command = apply_filters)
		apply_fil_btn.place(relx = 0.89, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		cancel_fil_btn = ttk.Button(main_window, text = "Cancel", command = cancel_filter)
		cancel_fil_btn.place(relx = 0.76, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		reset_fil_btn = ttk.Button(main_window, text = "Reset", command = reset_filter)
		reset_fil_btn.place(relx = 0.89, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
	elif (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		def apply_bulk_filter():
			operations(5, current_stack['Folder'])		
		def cancel_bulk_filter():
			folder_label_filter.destroy()
			folder_name_label_filter.destroy()
			apply_filter_btn.destroy()
			cancel_filter_btn.destroy()
		folder_label_filter = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label_filter.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label_filter = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label_filter.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		apply_filter_btn = ttk.Button(main_window, text = "Apply Mask", command = apply_bulk_filter)
		apply_filter_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_filter_btn = ttk.Button(main_window, text = "Cancel", command = cancel_bulk_filter)
		cancel_filter_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
##################################################################################################################
def colormap():
	color_datatypes = ['viridis', 'plasma', 'inferno', 'magma', 'cividis']
	message_color = ("Note: For Colormap operation\n'png' format would works well")	
	def cancel_color():
		filter_selection.destroy()
		combo_box_map.destroy()
		mask_map_btn.destroy()
		apply_map_btn.destroy()
		cancel_map_btn.destroy()
		reset_map_btn.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		normal()
	def reset_color():
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		src_color = cv2.imread("Work_Dir/img.png")
		cv2.imwrite("Work_Dir/img.png", src_color)
		disp_image("Work_Dir/img.png")
	def set_color():
		color_apply = variables_colormap.get()
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['G1']= 'Colormap'
		ws['G2'] = color_apply
		book.save('Support/log.xlsx')
		message_box("Done", "Mask setting are saved")
	def apply_colormap():
		def apply_action(choose_cl):
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			frame_dummy = cv2.imread("Work_Dir/img.png",0)
			colormap = plt.get_cmap(choose_cl)
			heatmap = (colormap(frame_dummy) * 2**16).astype(np.uint16)[:,:,:3]
			img_dummy = cv2.cvtColor(heatmap, cv2.COLOR_RGB2BGR)
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		color_apply = variables_colormap.get()
		if color_apply == "":
			message_box("Error", "No colormap is selected")
		elif color_apply == "viridis":
			apply_action("viridis")
		elif color_apply == "plasma":
			apply_action("plasma")
		elif color_apply == "inferno":
			apply_action("inferno")
		elif color_apply == "magma":
			apply_action("magma")
		elif color_apply == "cividis":
			apply_action("cividis")	
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		current_stack["Adjustments"] = "True"
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		filter_selection = ttk.Label(main_window, text = "Select Colormap ")
		filter_selection.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		global variables_colormap
		variables_colormap = StringVar()
		combo_box_map = ttk.Combobox(main_window,textvariable =variables_colormap,font = ('arial',8), values = color_datatypes)
		combo_box_map.place(relx = 0.86, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		mask_map_btn = ttk.Button(main_window, text = "Set Mask", command = set_color)
		mask_map_btn.place(relx = 0.76, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		apply_map_btn = ttk.Button(main_window, text = "Apply", command = apply_colormap)
		apply_map_btn.place(relx = 0.89, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		cancel_map_btn = ttk.Button(main_window, text = "Cancel", command = cancel_color)
		cancel_map_btn.place(relx = 0.76, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		reset_map_btn = ttk.Button(main_window, text = "Reset", command = reset_color)
		reset_map_btn.place(relx = 0.89, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
	elif (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		def apply_bulk_color():
			operations(6, current_stack['Folder'])		
		def cancel_bulk_color():
			folder_label_color.destroy()
			folder_name_label_color.destroy()
			apply_color_btn.destroy()
			cancel_color_btn.destroy()
			message_label_color.destroy()
			text_head_adj = ttk.Label(main_window)
			text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		folder_label_color = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label_color.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label_color = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label_color.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		apply_color_btn = ttk.Button(main_window, text = "Apply Mask", command = apply_bulk_color)
		apply_color_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_color_btn = ttk.Button(main_window, text = "Cancel", command = cancel_bulk_color)
		cancel_color_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		message_label_color = Label(main_window, text = message_color)
		message_label_color.place(relx = 0.75, rely = 0.84, relheight = 0.05, relwidth = 0.24)
####################################################################################################################
def blur():
	def apply_blur():
		blur_val = s3.get()
		blur_val= float(blur_val)
		blur_val = round(blur_val)
		blur_value_label = ttk.Label(main_window, text = str(blur_val))
		blur_value_label.place(relx = 0.93, rely = 0.63, relheight = 0.03, relwidth = 0.05)
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		blur_state = (blur_val, blur_val)
		frame_dummy = cv2.imread("Work_Dir/img.png")
		try:
			img_dummy = cv2.blur(frame_dummy, blur_state)
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		except:
			pass
	def reset_blur():
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		src_blur = cv2.imread("Work_Dir/img.png")
		s3.set(0)
		blur_val = s3.get()
		blur_val= float(blur_val)
		blur_val = round(blur_val)
		blur_value_label = ttk.Label(main_window, text = str(blur_val))
		blur_value_label.place(relx = 0.93, rely = 0.63, relheight = 0.03, relwidth = 0.05)
		cv2.imwrite("Work_Dir/img.png", src_blur)
		disp_image("Work_Dir/img.png")
	def cancel_blur():
		blur_label.destroy()
		mask_blur_btn.destroy()
		apply_blur_btn.destroy()
		cancel_blur_btn.destroy()
		reset_blur_btn.destroy()
		s3.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		#get_dummy("Support/dummy.png", "Work_Dir/img.png")
		#src_blur = cv2.imread("Work_Dir/img.png")
		#cv2.imwrite("Work_Dir/img.png", src_blur)
		#disp_image("Work_Dir/img.png")
		#src_blur = cv2.imread("Work_Dir/img.png")
		normal()
	def set_blur():
		blur_val = s3.get()
		blur_val= float(blur_val)
		blur_val = round(blur_val)
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['H1']= 'Blur'
		ws['H2'] = blur_val
		book.save('Support/log.xlsx')
		message_box("Done", "Mask setting are saved")
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		current_stack["Adjustments"] = "True"
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		s3 = ttk.Scale(main_window,from_ = 0, to = 50, orient = HORIZONTAL)
		s3.set(0) 
		s3.place(relx = 0.86, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		blur_label = ttk.Label(main_window, text = "Blur amount: ")
		blur_label.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.08)
		mask_blur_btn = ttk.Button(main_window, text = "Set Mask", command= set_blur)
		mask_blur_btn.place(relx = 0.76, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		apply_blur_btn = ttk.Button(main_window, text = "Apply", command = apply_blur)
		apply_blur_btn.place(relx = 0.89, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		cancel_blur_btn = ttk.Button(main_window, text = "Cancel", command = cancel_blur)
		cancel_blur_btn.place(relx = 0.76, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		reset_blur_btn = ttk.Button(main_window, text = "Reset", command = reset_blur)
		reset_blur_btn.place(relx = 0.89, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
	elif (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		def apply_bulk_blur():
			operations(7, current_stack['Folder'])		
		def cancel_bulk_blur():
			folder_label_blur.destroy()
			folder_name_label_blur.destroy()
			apply_blur_btn.destroy()
			cancel_blur_btn.destroy()
			text_head_adj = ttk.Label(main_window)
			text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		folder_label_blur = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label_blur.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label_blur = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label_blur.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		apply_blur_btn = ttk.Button(main_window, text = "Apply Mask", command = apply_bulk_blur)
		apply_blur_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_blur_btn = ttk.Button(main_window, text = "Cancel", command = cancel_bulk_blur)
		cancel_blur_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
###################################################################################################################
def flipping():
	def set_flip_mask():
		flip_apply = variables_flip.get()
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['I1']= 'Flip'
		ws['I2'] = flip_apply
		book.save('Support/log.xlsx')
		message_box("Done", "Mask setting are saved")
	def cancel_flip():
		mirror_selection.destroy()
		combo_box_flip.destroy()
		mask_flip_btn.destroy()
		apply_flip_btn.destroy()
		cancel_flip_btn.destroy()
		reset_flip_btn.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		normal()
	def apply_flip():
		def apply_action_flip(choose_flip):
			get_dummy("Support/dummy.png", "Work_Dir/img.png")
			frame_dummy = cv2.imread("Work_Dir/img.png")
			img_dummy = cv2.flip(frame_dummy, choose_flip)
			cv2.imwrite("Work_Dir/img.png", img_dummy)
			disp_image("Work_Dir/img.png")
		flip_apply = variables_flip.get()
		if flip_apply == "":
			message_box("Error", "No flip is selected")
		elif flip_apply == "Flip vertical":
			apply_action_flip(0)
		elif flip_apply == "Flip horizontal":
			apply_action_flip(1)
		elif flip_apply == "Flip horizontal & vertical":
			apply_action_flip(-1)
		else:
			message_box("Error", "Unknown option selected")
	def reset_flip():
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		src_flip = cv2.imread("Work_Dir/img.png")
		cv2.imwrite("Work_Dir/img.png", src_flip)
		disp_image("Work_Dir/img.png")
	flip_datatypes = ["Flip vertical", "Flip horizontal", "Flip horizontal & vertical"]
	if (current_stack["stack"] == ""):
		message_box("Error", "No image is selected")
	elif (current_stack["stack"] == "Single"):
		current_stack["Adjustments"] = "True"
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		mirror_selection = ttk.Label(main_window, text = "Select Flip Type: ")
		mirror_selection.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		global variables_flip
		variables_flip = StringVar()
		combo_box_flip = ttk.Combobox(main_window,textvariable =variables_flip,font = ('arial',7), values = flip_datatypes)
		combo_box_flip.place(relx = 0.855, rely = 0.6, relheight = 0.03, relwidth = 0.13)
		mask_flip_btn = ttk.Button(main_window, text = "Set Mask", command = set_flip_mask)
		mask_flip_btn.place(relx = 0.76, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		apply_flip_btn = ttk.Button(main_window, text = "Apply", command = apply_flip)
		apply_flip_btn.place(relx = 0.89, rely = 0.8, relheight = 0.03, relwidth = 0.09)
		cancel_flip_btn = ttk.Button(main_window, text = "Cancel", command = cancel_flip)
		cancel_flip_btn.place(relx = 0.76, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		reset_flip_btn = ttk.Button(main_window, text = "Reset", command = reset_flip)
		reset_flip_btn.place(relx = 0.89, rely = 0.84, relheight = 0.03, relwidth = 0.09)
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
	elif (current_stack["stack"] == "Bulk"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		def apply_bulk_flip():
			operations(8, current_stack['Folder'])		
		def cancel_bulk_flip():
			folder_label_flip.destroy()
			folder_name_label_flip.destroy()
			apply_flip_btn.destroy()
			cancel_flip_btn.destroy()
			text_head_adj = ttk.Label(main_window)
			text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		folder_label_flip = ttk.Label(main_window, text = "Selected Folder: ")
		folder_label_flip.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		basename =  os.path.basename(current_stack['Folder'])
		folder_name_label_flip = ttk.Label(main_window, text = "["+str(basename)+"]")
		folder_name_label_flip.place(relx = 0.85, rely = 0.6, relheight = 0.03, relwidth = 0.1)
		apply_flip_btn = ttk.Button(main_window, text = "Apply Mask", command = apply_bulk_flip)
		apply_flip_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_flip_btn = ttk.Button(main_window, text = "Cancel", command = cancel_bulk_flip)
		cancel_flip_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)

#####################################################################################################################
def camera_settings():
	def save_cam_settings():
		get_cam_id = variables_cam.get()
		if get_cam_id == "":
			message_box("Error", "Selection is empty")
		else:
			book = openpyxl.load_workbook('Support/log.xlsx')
			ws = book['Sheet1']
			ws['J1']= 'CAM'
			ws['J2'] = get_cam_id
			book.save('Support/log.xlsx')
			message_box("Done", "Camera setting are saved")
	def cancel_cam_settings():
		combo_box_cam.destroy()
		cam_name.destroy()
		save_cam_set.destroy()
		cancel_cam_set.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)

	cam_datatypes = ["Device 0", "Device 1", "Device 2"]
	message_box("Notify", "Closing working area")
	side_para("", "n")
	text_head_adj = ttk.Label(main_window)
	text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	try:
		disp_label.destroy()
	except:
		pass
	global variables_cam
	variables_cam = StringVar()
	combo_box_cam = ttk.Combobox(main_window,textvariable =variables_cam,font = ('arial',8), values = cam_datatypes)
	combo_box_cam.place(relx = 0.86, rely = 0.6, relheight = 0.03, relwidth = 0.1)
	cam_name = ttk.Label(main_window, text = "Select Camera:")
	cam_name.place(relx = 0.76, rely = 0.6, relheight = 0.03, relwidth = 0.1)
	save_cam_set = ttk.Button(main_window, text = "Save", command = save_cam_settings)
	save_cam_set.place(relx = 0.8, rely = 0.7, relheight = 0.03, relwidth = 0.05)
	cancel_cam_set = ttk.Button(main_window, text = "Close", command = cancel_cam_settings)
	cancel_cam_set.place(relx = 0.87, rely = 0.7, relheight = 0.03, relwidth = 0.05)
def capture_video_feed():
	try:
		disp_label.destroy()
	except:
		pass
	def optical_cam():
		global opt_cam, throw_label, cap
		opt_cam = Label(main_window)
		opt_cam.place(relx = 0.09, rely = 0.15, relheight = 0.7, relwidth = 0.6)
		df = pd.read_excel("Support/log.xlsx")
		text_cam = df.iloc[0,9]
		#print(text_cam)
		if (text_cam == "Device 0"):
			cap = cv2.VideoCapture(0)
		elif(text_cam == "Device 1"):
			cap = cv2.VideoCapture(1)
		elif(text_cam == "Device 2"):
			cap = cv2.VideoCapture(2)
		def show_frame():
			global frame
			ret, frame = cap.read()
			frame = cv2.flip(frame, 1)
			try:
				rgb_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
				img = PIL.Image.fromarray(rgb_img)
				imgtk = ImageTk.PhotoImage(image=img)
				opt_cam.imgtk = imgtk
				opt_cam.configure(image=imgtk)
				opt_cam.after(180, show_frame)
			except:
				pass
		show_frame()
	def cancel_cam():
		cap_btn.destroy()
		cancel_cap_btn.destroy()
		opt_cam.destroy()
		configcap_btn.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		normal()
		disp_label.destroy()
		try:
			cap.release()
		except:
			pass
	def cancel_captured():
		retake_btn.destroy()
		disp_label.destroy()
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		capture_video_feed()
	def capture_cam():
		try:
			cap.release()
		except:
			pass
		global retake_btn
		get_dummy("Support/dummy.png", "Work_Dir/img.png")
		save_photo = cv2.imwrite("Work_dir/img.png", frame)
		disp_image("Work_Dir/img.png")
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		retake_btn = ttk.Button(main_window, text = "Retake", command = cancel_captured)
		retake_btn.place(relx = 0.83, rely = 0.7, relheight = 0.03, relwidth = 0.08)

	if (current_stack["stack"] == "Single") or (current_stack["stack"] == "") or (current_stack["stack"] == "Bulk"):
		current_stack["Adjustments"] = "True"
		current_stack["stack"] = "Single"
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
		configcap_btn = ttk.Button(main_window, text = "Configure", command = camera_settings)
		configcap_btn.place(relx = 0.82, rely = 0.65, relheight = 0.03, relwidth = 0.09)
		cap_btn = ttk.Button(main_window, text = "Capture", command = capture_cam)
		cap_btn.place(relx = 0.76, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		cancel_cap_btn = ttk.Button(main_window, text = "Cancel",command = cancel_cam)
		cancel_cap_btn.place(relx = 0.89, rely = 0.7, relheight = 0.03, relwidth = 0.09)
		optical_cam()
		get_dummy("Work_Dir/img.png", "Support/dummy.png")
		side_para("", 'n')

	else:
		pass
#####################################################################################################################
def img_size(path):
	bytes_size = os.path.getsize(path)
	img_size = size(bytes_size)
	return img_size
def disp_image(path):
	global disp_label
	image = cv2.imread(path)
	resized_img = cv2.resize(image/255.0  , (630 , 500))
	cv2.imwrite('Work_Dir/img.png', 255*resized_img)
	disp_img = ImageTk.PhotoImage(PIL.Image.open('Work_Dir/img.png'))
	disp_label = Label(main_window, image = disp_img)
	disp_label.photo = disp_img
	disp_label.place(relx = 0.09, rely = 0.15, relheight = 0.7, relwidth = 0.6)
def cancel():
		side_para("", "n")
		try:
			os.remove("Work_Dir/original_copy.png")
		except:
			pass
def create_new_image():
	if (current_stack["Adjustments"] == "True"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	else:
		pass
	current_stack["Adjustments"] = "False"
	current_stack["stack"] = "Single"
	def create():
		def image(color, height, width):
			if color == 'Black':
				image = np.zeros(shape=[int(height), int(width), 3], dtype=np.uint8)
				cv2.imwrite('Work_Dir/img.png', image)
				cv2.imwrite('Work_Dir/original_copy.png', image)
			elif color == "White":
				image = 255 * np.ones(shape=[int(height), int(width), 3], dtype=np.uint8)
				cv2.imwrite('Work_Dir/img.png', image)
				cv2.imwrite('Work_Dir/original_copy.png', image)
		global disp_label
		variables = variables_bg.get()
		height  = height_val.get()
		width = width_val.get()
		if (variables == "") or (height == "") or (width == ""):
			message_box('Empty', 'Details are missing')
		elif (variables != "Black") and (variables != "White"):
			message_box("Error", "Unknown Background")
			try:
				disp_label.destroy()
			except:
				pass
		else: 
			height = float(height)
			width = float(width)
			book = openpyxl.load_workbook('Support/log.xlsx')
			ws = book['Sheet1']
			ws['A1']= 'NEW'
			ws['A2'] = variables
			ws['A3'] = round(height)
			ws['A4'] = round(width)
			book.save('Support/log.xlsx')
			cancel()	
			image(variables, height, width)
			path = 'Work_Dir/img.png'
			side_para("Background: \n " + (variables)+ "\n Height \n " + str(height) +"\n Width \n " + str(width)+"\n Total Size: "+str(img_size(path)), 'nw')
			disp_image(path)
	side_para("Background:\n "" \n Height:\n""\n Width", 'nw')
	bg_color = ["Black", "White"]
	global variable_bg
	variables_bg = StringVar()
	combo_box = ttk.Combobox(main_window,textvariable = variables_bg,font = ('arial',10), values = bg_color)
	combo_box.insert(0, new_color)
	combo_box.place(relx = 0.85, rely = 0.14, relheight = 0.03, relwidth = 0.1)
	global height_val
	height_val = StringVar()
	height_entry = Entry(main_window,textvariable = height_val)
	height_entry.insert(0, new_height)
	height_entry.place(relx = 0.85, rely = 0.19, relheight = 0.03, relwidth = 0.1)
	global width_val
	width_val = StringVar()
	width_entry = Entry(main_window,textvariable = width_val)
	width_entry.insert(0, new_width)
	width_entry.place(relx = 0.85, rely = 0.24, relheight = 0.03, relwidth = 0.1)
	create_btn = ttk.Button(main_window,text = "Create", command = create)
	create_btn.place(relx = 0.76, rely = 0.4, relheight = 0.03, relwidth = 0.1)
	cancel_btn = ttk.Button(main_window,text = "Cancel", command = cancel)
	cancel_btn.place(relx = 0.88, rely = 0.4, relheight = 0.03, relwidth = 0.1)
def open_img():
	if (current_stack["Adjustments"] == "True"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	else:
		pass
	current_stack["stack"] = "Single"
	current_stack["Adjustments"] = "False"
	try:
		global filename
		filename = filedialog.askopenfilename()
		global get_img
		get_img = filename
		basename =  os.path.basename(get_img)
		image = cv2.imread(get_img)
		global h, w
		h, w, c = image.shape
		book = openpyxl.load_workbook('Support/log.xlsx')
		ws = book['Sheet1']
		ws['A1']= 'NEW'
		ws['A2'] = 'N/A'
		ws['A3'] = float(h)
		ws['A4'] = float(w)
		book.save('Support/log.xlsx')
		cv2.imwrite('Work_Dir/original_copy.png', image)
		disp_image(get_img)
		side_para("Custom Selected Image\n"+str(basename)+"\nHeight \n " + str(h) +"\nWidth \n " + str(w)+"\nTotal Size: "+str(img_size(get_img)), 'nw')
	except:
		pass
def directory_selector():
		filename_dir = filedialog.askdirectory()
		global get_dir
		get_dir = filename_dir
		return get_dir
def open_bulk_imgs():
	if (current_stack["Adjustments"] == "True"):
		text_head_adj = ttk.Label(main_window)
		text_head_adj.place(relx = 0.75, rely = 0.55, relheight = 0.35, relwidth = 0.24)
	else:
		pass
	try:
		def remove():
			side_para("", "n")
			current_stack["Folder"] = ""
			current_stack["stack"] = ""
			label.destroy()
		directory_selector()
		current_stack["stack"] = "Bulk"
		current_stack["Folder"] = get_dir
		def get_size(get_dir):
			total_size = 0
			for dirpath, dirnames, filenames in os.walk(get_dir):
				for f in filenames:
					fp = os.path.join(dirpath, f)
					if not os.path.islink(fp):
						total_size += os.path.getsize(fp)
			t_size = size(total_size)
			return t_size
		folder_name = os.path.basename(get_dir)
		if folder_name == "":
			current_stack["stack"] = "Single"
			pass
		else:
			side_para("Custom Selected Folder\nFolder name: "+folder_name+" \nTotal Size: "+str(get_size(get_dir)), 'nw')
			icon_folder = ImageTk.PhotoImage(PIL.Image.open(icon('folder.png')))
			label =Label(main_window,image = icon_folder)
			label.photo = icon_folder
			label.place(relx = 0.35, rely = 0.45, relheight = 0.08, relwidth = 0.07)
			remove_btn = ttk.Button(main_window,text = "Remove", command = remove)
			remove_btn.place(relx = 0.832, rely = 0.47, relheight = 0.03, relwidth = 0.07)
			disp_label.destroy()
	except:
		pass
def save():
	if (current_stack["stack"] == "Single"):
		db()
		save_types = ['.jpg','.png', '.bmp','.tiff']
		def path_label_show(get_img_dir):
			label_path =ttk. Label(main_window, text = get_img_dir)
			label_path.place(relx = 0.85, rely = 0.35, relheight = 0.03, relwidth = 0.06)
		def directory_selection():
			filename_dir = filedialog.askdirectory()
			global get_img_dir
			get_img_dir = filename_dir
			current_stack["Folder"] = get_img_dir
			folder_name = os.path.basename(get_img_dir)
			path_label_show(folder_name)
		def copies():
			global copy_val, copy_entry
			copy_val = StringVar()
			copy_entry = Entry(main_window,textvariable = copy_val)
			copy_entry.place(relx = 0.89, rely = 0.44, relheight = 0.02, relwidth = 0.05)
		def c_1():
			copies()
			CheckVar2.set(0)
			get_val_check = CheckVar1.get()
		def c_2():
			try:
				copy_entry.destroy()
			except:
				pass
			CheckVar1.set(0)
			get_val_check = CheckVar2.get()
		def save_image():
			img_path = "Work_Dir/img.png"
			get_val_check_1 = CheckVar1.get()
			get_val_check_2 = CheckVar2.get()
			height  = height_val.get()
			width = width_val.get()
			img_name = (name_val).get()
			save_type = (variables_savetype).get()
			height = float(height)
			width = float(width)
			img = cv2.imread(img_path)
			resized_img = cv2.resize(img , (round(width), round(height)))
			save_path = (str(get_img_dir)+'/')+str(img_name)+str(save_type)
			if  (get_val_check_1) == 1:
				#print('multipes')
				if (height == "") or(width == "") or (img_name == "") or (save_type == ""):
					message_box('Error', 'Details are missing')
				else:
					copy_value = (copy_val).get()
					if (copy_value == ""):
						message_box('Error', 'Copy value is missing')
					else:
						if (current_stack["Adjustments"] == "True"):
							copy_value = int(copy_value)
							img_name = (name_val).get()
							save_type =(variables_savetype).get()
							for i in range(copy_value):
								save_path = (str(get_img_dir)+'/')+str(img_name)+str(i)+str(save_type)
								cv2.imwrite(save_path, resized_img)
							message_box('Saved', 'Images have been saved')
						else:
							try:
								if (h ==round(height)) and (w == round(width)):
									#print("yes")
									copy_value = int(copy_value)
									img_name = (name_val).get()
									save_type =(variables_savetype).get()
									my_image = cv2.imread("Work_Dir/original_copy.png")
									for i in range(copy_value):
										save_path = (str(get_img_dir)+'/')+str(img_name)+str(i)+str(save_type)
										cv2.imwrite(save_path, my_image)
									message_box('Saved', 'Images have been saved')
								else:
									#print("no")
									copy_value = int(copy_value)
									img_name = (name_val).get()
									save_type =(variables_savetype).get()
									for i in range(copy_value):
										save_path = (str(get_img_dir)+'/')+str(img_name)+str(i)+str(save_type)
										cv2.imwrite(save_path, resized_img)
									message_box('Saved', 'Images have been saved')
							except:
								copy_value = int(copy_value)
								img_name = (name_val).get()
								save_type =(variables_savetype).get()
								for i in range(copy_value):
									save_path = (str(get_img_dir)+'/')+str(img_name)+str(i)+str(save_type)
									cv2.imwrite(save_path, resized_img)
								message_box('Saved', 'Images have been saved')
			elif (get_val_check_2) == 1:
				if (height == "") or(width == "") or (img_name == "") or (save_type == ""):
					message_box('Error', 'Details are missing')
				else:
					if (current_stack["Adjustments"] == "True"):
						cv2.imwrite(save_path, resized_img)
						message_box('Saved', str(img_name)+' has been saved')
					else:
						try:
							if (h ==round(height)) and (w == round(width)):
								my_image = cv2.imread("Work_Dir/original_copy.png")
								cv2.imwrite(save_path, my_image)
								message_box('Saved', str(img_name)+' has been saved')
							else:
								cv2.imwrite(save_path, resized_img)
								message_box('Saved', str(img_name)+' has been saved')
						except:
							cv2.imwrite(save_path, resized_img)
							message_box('Saved', str(img_name)+' has been saved')
			
		side_para("Name:\n""\nType:\n""\nHeight:\n""\nWidth\n""\nFolder Name:\n""\nGenerate Multiple Copies", 'nw')
		path_label_show("")
		browse_btn = ttk.Button(main_window,text = "Browse",command = directory_selection)
		browse_btn.place(relx = 0.925, rely = 0.35, relheight = 0.03, relwidth = 0.06)
		save_btn = ttk.Button(main_window,text = "Save", command = save_image)
		save_btn.place(relx = 0.76, rely = 0.47, relheight = 0.03, relwidth = 0.1)
		cancel_btn = ttk.Button(main_window,text = "Cancel", command = cancel)
		cancel_btn.place(relx = 0.88, rely = 0.47, relheight = 0.03, relwidth = 0.1)
		C1 = ttk.Checkbutton(main_window, text = "Yes", variable = CheckVar1,onvalue = 1, offvalue = 0,cursor = "dot",command = c_1)
		C2 = ttk.Checkbutton(main_window, text = "No", variable = CheckVar2, onvalue = 1, offvalue = 0,cursor = "dot", command = c_2)
		CheckVar2.set(1)
		C1.place(relx = 0.76, rely = 0.44, relheight = 0.02, relwidth = 0.05)
		C2.place(relx = 0.83, rely = 0.44, relheight = 0.02, relwidth = 0.05)
		global height_val
		height_val = StringVar()
		height_entry = Entry(main_window,textvariable = height_val)
		height_entry.insert(0, new_height)
		height_entry.place(relx = 0.85, rely = 0.24, relheight = 0.03, relwidth = 0.1)
		global width_val
		width_val = StringVar()
		width_entry = Entry(main_window,textvariable = width_val)
		width_entry.insert(0, new_width)
		width_entry.place(relx = 0.85, rely = 0.3, relheight = 0.03, relwidth = 0.1)
		global name_val
		name_val = StringVar()
		name_entry = Entry(main_window,textvariable = name_val)
		name_entry.place(relx = 0.85, rely = 0.14, relheight = 0.03, relwidth = 0.1)
		global variables_savetype
		variables_savetype = StringVar()
		combo_box = ttk.Combobox(main_window,textvariable = variables_savetype,font = ('arial',10), values = save_types)
		combo_box.place(relx = 0.85, rely = 0.19, relheight = 0.03, relwidth = 0.1)
	elif(current_stack["stack"] == "Bulk"):
		message_box("False selection","You have selected bulk images\nMask actions are required" )
	else:
		message_box("Error", "No image is selected")
def abt():
	message_box("About","MULTIPURPOSE BATCH IMAGE PROCESSING APPLICATION\n\nVersion: 1.0.0\n\nMultipurpose batch image processing application is a free and open source bulk image processing portal, programmed in python3.\n\nThis application will help in applying image operations for image batches in a directory")
def credits():
	message_box("Application credits","Builded by: Manjunathan S\n\nProject Timeline: NOV 2020 - DEC 2020\n\nContact: manjunathsg407@gmail.com\n\nFeel Free To Contact Me!")
def screen():
	global main_window
	global side_para
	global adj_para
	def text_heading(p):
		if p == 1:
			return "Image Parameters"
		elif p == 2:
			return "Image Adjustments"
	main_window = Frame(window, bg = "gray15")
	main_window.pack(fill = BOTH,expand = True)
	display_screen =ttk.Label(main_window)
	display_screen.place(relx = 0.04, rely = 0.1, relheight = 0.8, relwidth = 0.7)
	def side_para(txt, anc):
		side_parameters = ttk.Label(main_window, text = txt, anchor = anc,font =('bahnschrift',12))
		side_parameters.place(relx = 0.75, rely = 0.135, relheight = 0.37, relwidth = 0.24)
	side_para("", 'n')
	def adj_para(txt, anc):
		adj_parameters = ttk.Label(main_window, text = txt,anchor = anc,font =('bahnschrift',12))
		adj_parameters.place(relx = 0.75, rely = 0.545, relheight = 0.355, relwidth = 0.24)
	adj_para("", 'n')
	text_head_para = Label(main_window, text = text_heading(1), font =('arial',10,'bold'),bg = "gray75")
	text_head_para.place(relx = 0.75, rely = 0.1, relheight = 0.03, relwidth = 0.24)
	text_head_adj = Label(main_window, text = text_heading(2), font =('arial',10,'bold'), bg = "gray75")
	text_head_adj.place(relx = 0.75, rely = 0.51, relheight = 0.03, relwidth = 0.24)
	menu_bar = Menu(main_window)
	window.config(menu = menu_bar)
	submenu_1 = Menu(menu_bar,tearoff = 0)
	submenu_2 = Menu(menu_bar,tearoff = 0)
	submenu_3 = Menu(menu_bar,tearoff = 0)
	submenu_4 = Menu(menu_bar,tearoff = 0)
	menu_bar.add_cascade(label = "File", menu = submenu_1)
	menu_bar.add_cascade(label = "Edit", menu = submenu_2)
	menu_bar.add_cascade(label = "Image", menu = submenu_3)
	menu_bar.add_cascade(label = "Help", menu = submenu_4)
	submenu_1.add_command(label ="Create New Image", command = create_new_image)
	submenu_1.add_command(label ="Open Image", command = open_img)
	submenu_1.add_command(label ="Select Folder", command = open_bulk_imgs)
	submenu_1.add_command(label ="Save as...", command = save)
	submenu_1.add_command(label ="Close")
	submenu_1.add_command(label ="Exit", command = exit)
	submenu_2.add_command(label ="Image Corrections", command = img_corrections)
	submenu_2.add_command(label ="Resize Images", command = resizer)
	submenu_2.add_command(label ="Image Orientation", command = image_orient)
	submenu_2.add_command(label ="Rename Images", command = rename_images)
	submenu_2.add_command(label ="Bitwise Operations", command = bitwise)
	submenu_2.add_command(label ="Image Augmentation", command = data_augment)
	submenu_3.add_command(label ="Histogram", command = histogram)
	submenu_3.add_command(label ="Filter Tools", command = filter_tools)
	submenu_3.add_command(label ="Colormap", command = colormap)
	submenu_3.add_command(label ="Blur", command = blur)
	submenu_3.add_command(label ="Flip", command = flipping)
	submenu_3.add_command(label ="Camera", command = capture_video_feed)
	submenu_4.add_command(label ="About", command = abt)
	submenu_4.add_command(label ="Credits", command = credits)
screen()
window.protocol("WM_DELETE_WINDOW", exit)
window.mainloop()
